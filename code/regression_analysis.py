"""
Baseline Gut Multi-Omics Regression Analysis
=============================================

This script implements the core machine-learning pipeline used in the
manuscript "Baseline Gut Multi-Omics Profiles Predict Donor-Specific
Metabolomic Responses to Dietary Substrates: A Regression-Based Machine
Learning Approach".

Pipeline overview
------------------
1. Load raw 16S rRNA (taxonomy), metabolome, and metadata tables.
2. Restrict to donors with paired Control/FOS/Laminarin/Onion samples.
3. Compute a PC1-based metabolome "response score" (treatment vs. matched
   control) for each donor and substrate.
4. Build three baseline predictor sets (Taxonomy-only, Metabolome-only,
   Combined) using each donor's pre-treatment (Control) profile.
5. Evaluate ElasticNet, Random Forest, and SVR (linear kernel) with
   Leave-One-Donor-Out cross-validation (LODO-CV).
6. Select the best-performing input/model combination (by Spearman r).
7. Compute permutation importance for (a) the overall baseline model and
   (b) substrate-specific models (FOS, Laminarin, Onion).
8. Export all performance metrics, predictions, and importance scores to
   an Excel workbook.

Required input files (see ../data/)
------------------------------------
  Merge_metadata.txt
  feature-count_clr_transposed.tsv
  metabolite_common_log2_zscore_transposed.tsv

Output
------
  ../results/regression_analysis_output.xlsx
"""

import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

from sklearn.base import clone
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.impute import SimpleImputer
from sklearn.linear_model import ElasticNetCV
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.model_selection import LeaveOneGroupOut
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.inspection import permutation_importance
from sklearn.decomposition import PCA

from scipy.stats import spearmanr, pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# =====================================================================
# 1. SETTINGS
# =====================================================================
ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data"
RESULTS_DIR = ROOT_DIR / "results"
RESULTS_DIR.mkdir(exist_ok=True)

META_FILE = DATA_DIR / "Merge_metadata.txt"
MICRO_FILE = DATA_DIR / "feature-count_clr_transposed.tsv"
METAB_FILE = DATA_DIR / "metabolite_common_log2_zscore_transposed.tsv"

OUTPUT_EXCEL = RESULTS_DIR / "regression_analysis_output.xlsx"

GROUP_ORDER = ["Control", "FOS", "Onion", "Laminarin"]
TREATMENTS = ["FOS", "Laminarin", "Onion"]
INPUT_NAMES = ["Taxonomy", "Metabolome", "Combined"]

N_PERMUTATION_REPEATS = 30
RANDOM_STATE = 42

MODELS = {
    "ElasticNet": ElasticNetCV(
        l1_ratio=[0.1, 0.3, 0.5, 0.7, 0.9, 1.0],
        alphas=np.logspace(-3, 1, 50),
        cv=5, max_iter=10000, random_state=RANDOM_STATE,
    ),
    "RandomForest": RandomForestRegressor(
        n_estimators=500, min_samples_leaf=2, random_state=RANDOM_STATE,
    ),
    "SVR": SVR(kernel="linear", C=1.0, epsilon=0.1),
}
MODEL_NAMES = list(MODELS.keys())


# =====================================================================
# 2. HELPER FUNCTIONS
# =====================================================================
def clean_feature_name(name):
    """Convert a raw ASV/feature ID into a human-readable taxon/metabolite name."""
    if name is None:
        return "Unknown feature"
    name = re.sub(r"(_tax|_met)$", "", str(name).strip())
    if not name or name.lower() == "nan":
        return "Unknown feature"

    # Taxonomy strings (e.g. "d__Bacteria;p__...;g__...")
    if any(tag in name for tag in ["d__", "p__", "c__", "o__", "f__", "g__", "s__"]):
        parts = [p.strip() for p in name.split(";") if p.strip()]
        rank_map = {}
        for p in parts:
            if "__" in p:
                rank, value = p.split("__", 1)
                rank_map[rank] = value.strip()
        for rank in ["g", "f", "o", "c", "p", "s"]:
            value = rank_map.get(rank, "")
            if value and value not in ["_", "uncultured", "unclassified", "unknown", "Unassigned", ""]:
                return value.replace("_", " ")
        if parts:
            last = re.sub(r"^[a-z]__", "", parts[-1]).strip()
            if last:
                return last.replace("_", " ")
        return "Unclassified taxon"

    return name.replace("_", " ").strip() or "Unknown feature"


def style_worksheet(ws):
    """Apply consistent header/column styling to an openpyxl worksheet."""
    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    centered = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = centered
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=4)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)


def add_dataframe_sheet(workbook, sheet_name, df):
    ws = workbook.create_sheet(sheet_name[:31])
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    style_worksheet(ws)


# =====================================================================
# 3. LOAD RAW DATA
# =====================================================================
def load_raw_data():
    print("[1/7] Loading raw data ...")
    meta = pd.read_csv(META_FILE, sep="\t").rename(columns={"#SampleID": "SampleID"})
    micro = pd.read_csv(MICRO_FILE, sep="\t")
    metab = pd.read_csv(METAB_FILE, sep="\t")

    if "sample_id" in micro.columns:
        micro = micro.rename(columns={"sample_id": "SampleID"})
    if "sample_id" in metab.columns:
        metab = metab.rename(columns={"sample_id": "SampleID"})

    return meta, micro, metab


# =====================================================================
# 4. SELECT DONORS WITH COMPLETE PAIRED SAMPLES
# =====================================================================
def get_complete_donors(meta, micro, metab):
    meta = meta[meta["Group"].isin(GROUP_ORDER)].copy()

    group_counts = meta.groupby("donor_id")["Group"].nunique()
    complete_donors = set(group_counts[group_counts == len(GROUP_ORDER)].index)
    complete_donors &= set(meta.loc[meta["SampleID"].isin(micro["SampleID"]), "donor_id"])
    complete_donors &= set(meta.loc[meta["SampleID"].isin(metab["SampleID"]), "donor_id"])
    complete_donors = sorted(complete_donors)

    meta_complete = meta[meta["donor_id"].isin(complete_donors)].copy()
    print(f"      n complete donors = {len(complete_donors)}")
    return meta_complete


# =====================================================================
# 5. METABOLOME RESPONSE SCORE (PC1-based delta)
# =====================================================================
def compute_response_scores(meta_complete, metab):
    print("[2/7] Computing PC1-based metabolome response scores ...")
    merged = meta_complete[["SampleID", "donor_id", "Group"]].merge(metab, on="SampleID", how="inner")
    feature_cols = [c for c in merged.columns if c not in ["SampleID", "donor_id", "Group"]]

    X = merged[feature_cols].apply(pd.to_numeric, errors="coerce")
    X = X.replace([np.inf, -np.inf], np.nan).fillna(0)

    pcs = PCA(n_components=2, random_state=RANDOM_STATE).fit_transform(StandardScaler().fit_transform(X))
    merged["PC1"] = pcs[:, 0]

    wide = merged.pivot(index="donor_id", columns="Group", values="PC1").reset_index()

    # Orient PC1 so that "Onion" responses are positive on average
    if (wide["Onion"] - wide["Control"]).mean() < 0:
        wide[GROUP_ORDER] *= -1

    for trt in TREATMENTS:
        wide[f"delta_{trt}"] = wide[trt] - wide["Control"]

    response_wide = wide[["donor_id"] + [f"delta_{t}" for t in TREATMENTS]].copy()
    response_wide["Overall_Response"] = response_wide[[f"delta_{t}" for t in TREATMENTS]].mean(axis=1)

    response_long = (
        response_wide[["donor_id"] + [f"delta_{t}" for t in TREATMENTS]]
        .melt(id_vars="donor_id", var_name="Treatment", value_name="Response")
    )
    response_long["Treatment"] = response_long["Treatment"].str.replace("delta_", "", regex=False)
    response_long = response_long.sort_values(["donor_id", "Treatment"]).reset_index(drop=True)

    return response_long


# =====================================================================
# 6. BUILD BASELINE (PRE-TREATMENT) PREDICTOR TABLES
# =====================================================================
def build_baseline_tables(meta_complete, micro, metab, response_long):
    control_meta = meta_complete[meta_complete["Group"] == "Control"][["SampleID", "donor_id"]].copy()

    def to_baseline(raw_df):
        df = control_meta.merge(raw_df, on="SampleID", how="inner").drop(columns=["SampleID"]).copy()
        feature_cols = [c for c in df.columns if c != "donor_id"]
        df[feature_cols] = df[feature_cols].apply(pd.to_numeric, errors="coerce")
        return df

    micro_baseline = to_baseline(micro)
    metab_baseline = to_baseline(metab)
    combined_baseline = pd.merge(
        micro_baseline, metab_baseline, on="donor_id", how="inner", suffixes=("_tax", "_met")
    )

    datasets = {
        "Taxonomy": response_long.merge(micro_baseline, on="donor_id", how="inner"),
        "Metabolome": response_long.merge(metab_baseline, on="donor_id", how="inner"),
        "Combined": response_long.merge(combined_baseline, on="donor_id", how="inner"),
    }
    return datasets


# =====================================================================
# 7. PREPROCESSING PIPELINES
# =====================================================================
def make_preprocessor_with_treatment(df):
    predictors = [c for c in df.columns if c not in ["donor_id", "Treatment", "Response"]]
    preprocessor = ColumnTransformer([
        ("num", Pipeline([("impute", SimpleImputer(strategy="median")),
                          ("scale", StandardScaler())]), predictors),
        ("cat", OneHotEncoder(drop="first", handle_unknown="ignore"), ["Treatment"]),
    ], remainder="drop")
    return preprocessor, predictors


def make_preprocessor_baseline_only(df):
    predictors = [c for c in df.columns if c not in ["donor_id", "Treatment", "Response"]]
    preprocessor = ColumnTransformer([
        ("num", Pipeline([("impute", SimpleImputer(strategy="median")),
                          ("scale", StandardScaler())]), predictors),
    ], remainder="drop")
    return preprocessor, predictors


# =====================================================================
# 8. LODO-CV EVALUATION (3 inputs x 3 models)
# =====================================================================
def evaluate_lodo_cv(df, input_name, logo):
    """Leave-One-Donor-Out CV for one input dataset across all models."""
    preprocessor, predictors = make_preprocessor_with_treatment(df)
    X = df[["Treatment"] + predictors].copy()
    y = df["Response"].values
    groups = df["donor_id"].values

    perf_rows = []
    predictions_by_model = {}

    for model_name, model in MODELS.items():
        y_true, y_pred, donors, treatments = [], [], [], []
        for train_idx, test_idx in logo.split(X, y, groups):
            pipe = Pipeline([("pre", preprocessor), ("model", clone(model))])
            pipe.fit(X.iloc[train_idx], y[train_idx])
            y_pred.extend(pipe.predict(X.iloc[test_idx]).tolist())
            y_true.extend(y[test_idx].tolist())
            donors.extend(df.iloc[test_idx]["donor_id"].tolist())
            treatments.extend(df.iloc[test_idx]["Treatment"].tolist())

        y_true, y_pred = np.array(y_true), np.array(y_pred)
        rho, rho_p = spearmanr(y_true, y_pred)
        r, r_p = pearsonr(y_true, y_pred)

        perf_rows.append({
            "Input": input_name, "Model": model_name, "N": len(y_true),
            "Spearman_r": rho, "Spearman_p": rho_p,
            "Pearson_r": r, "Pearson_p": r_p,
            "RMSE": np.sqrt(mean_squared_error(y_true, y_pred)),
            "R2": r2_score(y_true, y_pred),
        })
        predictions_by_model[model_name] = pd.DataFrame({
            "donor_id": donors, "Treatment": treatments,
            "Observed": y_true, "Predicted": y_pred,
            "Input": input_name, "Model": model_name,
        })

    return pd.DataFrame(perf_rows), predictions_by_model


# =====================================================================
# 9. PERMUTATION IMPORTANCE
# =====================================================================
def compute_permutation_importance(df, model_name, treatment=None):
    """Fit the full model on `df` and compute permutation importance.

    If `treatment` is given, the result table includes a "Treatment" column
    (used for substrate-specific importance).
    """
    preprocessor, predictors = make_preprocessor_baseline_only(df)
    X = df[predictors].copy()
    y = df["Response"].values

    pipe = Pipeline([("pre", preprocessor), ("model", clone(MODELS[model_name]))])
    pipe.fit(X, y)

    result = permutation_importance(
        pipe, X, y, scoring="neg_mean_squared_error",
        n_repeats=N_PERMUTATION_REPEATS, random_state=RANDOM_STATE,
    )

    importance_df = pd.DataFrame({
        "Feature": predictors,
        "Importance_mean": result.importances_mean,
        "Importance_std": result.importances_std,
    }).sort_values("Importance_mean", ascending=False).reset_index(drop=True)

    if treatment is not None:
        importance_df.insert(0, "Treatment", treatment)

    importance_df["Feature_clean"] = importance_df["Feature"].apply(clean_feature_name)
    importance_df["Feature_clean"] = importance_df["Feature_clean"].replace("", np.nan).fillna("Unclassified taxon")

    # Keep only the highest-importance row per cleaned feature name
    importance_df = importance_df.drop_duplicates(subset="Feature_clean", keep="first").reset_index(drop=True)
    return importance_df


def evaluate_substrate_specific(df, model_name, input_name, logo):
    """LODO-CV + permutation importance, run separately for each substrate."""
    perf_rows = []
    importance_by_treatment = {}

    for trt in TREATMENTS:
        sub = df[df["Treatment"] == trt].copy()
        preprocessor, predictors = make_preprocessor_baseline_only(sub)
        X = sub[predictors].copy()
        y = sub["Response"].values
        groups = sub["donor_id"].values

        y_true, y_pred = [], []
        for train_idx, test_idx in logo.split(X, y, groups):
            pipe = Pipeline([("pre", preprocessor), ("model", clone(MODELS[model_name]))])
            pipe.fit(X.iloc[train_idx], y[train_idx])
            y_pred.extend(pipe.predict(X.iloc[test_idx]).tolist())
            y_true.extend(y[test_idx].tolist())

        y_true, y_pred = np.array(y_true), np.array(y_pred)
        rho, rho_p = spearmanr(y_true, y_pred)
        r, r_p = pearsonr(y_true, y_pred)

        perf_rows.append({
            "Input": input_name, "Model": model_name, "Treatment": trt,
            "N": len(y_true), "Spearman_r": rho, "Spearman_p": rho_p,
            "Pearson_r": r, "Pearson_p": r_p,
            "RMSE": np.sqrt(mean_squared_error(y_true, y_pred)),
            "R2": r2_score(y_true, y_pred),
        })

        importance_by_treatment[trt] = compute_permutation_importance(sub, model_name, treatment=trt)

    return pd.DataFrame(perf_rows), importance_by_treatment


# =====================================================================
# 10. MAIN PIPELINE
# =====================================================================
def main():
    meta, micro, metab = load_raw_data()
    meta_complete = get_complete_donors(meta, micro, metab)
    response_long = compute_response_scores(meta_complete, metab)
    datasets = build_baseline_tables(meta_complete, micro, metab, response_long)

    # --- LODO-CV across 3 inputs x 3 models -------------------------
    print("[3/7] Running Leave-One-Donor-Out CV (3 inputs x 3 models) ...")
    logo = LeaveOneGroupOut()
    perf_frames, predictions = [], {}
    for input_name, df in datasets.items():
        perf_df, preds = evaluate_lodo_cv(df, input_name, logo)
        perf_frames.append(perf_df)
        predictions[input_name] = preds
    perf_df = pd.concat(perf_frames, ignore_index=True)

    # --- Select best input/model combination ------------------------
    print("[4/7] Selecting best-performing input/model combination ...")
    best_row = perf_df.sort_values("Spearman_r", ascending=False).iloc[0]
    best_input, best_model = best_row["Input"], best_row["Model"]
    print(f"      -> {best_input} / {best_model}  (Spearman r = {best_row['Spearman_r']:.3f})")

    best_predictions = predictions[best_input][best_model]
    best_dataset = datasets[best_input].copy()

    # --- Baseline (overall) permutation importance -------------------
    print("[5/7] Computing baseline permutation importance ...")
    baseline_importance = compute_permutation_importance(best_dataset, best_model)

    # --- Substrate-specific models + importance -----------------------
    print("[6/7] Running substrate-specific models and importance ...")
    ts_perf, ts_importance = evaluate_substrate_specific(best_dataset, best_model, best_input, logo)

    # --- Save everything to Excel -------------------------------------
    print("[7/7] Saving results ...")
    wb = Workbook()
    wb.remove(wb.active)
    add_dataframe_sheet(wb, "Model_Performance", perf_df)
    add_dataframe_sheet(wb, "BestModel_Predictions", best_predictions)
    add_dataframe_sheet(wb, "Importance_BaselineOnly", baseline_importance)
    add_dataframe_sheet(wb, "TreatmentSpecific_Perf", ts_perf)
    for trt in TREATMENTS:
        add_dataframe_sheet(wb, f"TS_Importance_{trt}", ts_importance[trt])
    wb.save(OUTPUT_EXCEL)
    print(f"      Saved: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
